import asyncio
import json
import ast
from datetime import datetime, timezone
from fastapi import APIRouter, Depends, HTTPException, WebSocket, WebSocketDisconnect, BackgroundTasks
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from sqlalchemy.orm import selectinload
from io import BytesIO
from fastapi.responses import StreamingResponse, JSONResponse

from app.db.database import get_db
from app.models.models import AnalysisJob, AnalysisResult, BehaviourSegment, VideoUpload
from app.schemas.schemas import (
    AnalysisJobCreate, AnalysisJobResponse, AnalysisResultResponse,
    ADOSCoreScores, HRIExtensionScores, ProfileSummary,
    FrameSummariesResponse, FrameSummaryResponse
)
from app.services.pipeline import AnalysisPipeline
from app.core.config import get_settings
from app.core.dependencies import get_current_user
from app.models.models import AnalysisJob, AnalysisResult, BehaviourSegment, VideoUpload, User
import structlog

logger = structlog.get_logger()
router = APIRouter(prefix="/analysis", tags=["analysis"])
settings = get_settings()

# Active WebSocket connections: {job_id: WebSocket}
_ws_connections: dict[str, WebSocket] = {}


@router.post("/start", response_model=AnalysisJobResponse)
async def start_analysis(
    payload: AnalysisJobCreate,
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Start a new analysis job for an uploaded video."""
    # Verify video exists
    result = await db.execute(select(VideoUpload).where(VideoUpload.id == payload.video_id))
    video = result.scalar_one_or_none()
    if not video:
        raise HTTPException(status_code=404, detail="Video not found")

    # Create job
    job = AnalysisJob(
        video_id=payload.video_id,
        status="queued",
        vlm_model=payload.vlm_model or settings.vlm_model,
    )
    db.add(job)
    await db.commit()
    await db.refresh(job)

    # Run pipeline in background
    background_tasks.add_task(
        run_pipeline_background,
        job_id=job.id,
        video_path=video.upload_path,
        vlm_model=job.vlm_model,
    )

    logger.info("analysis_started", job_id=job.id, video_id=payload.video_id)
    return AnalysisJobResponse.model_validate(job)


async def run_pipeline_background(job_id: str, video_path: str, vlm_model: str):
    """Background task: run the full 7-stage pipeline and persist results."""
    from app.db.database import AsyncSessionLocal

    async with AsyncSessionLocal() as db:
        # Mark as running
        result = await db.execute(select(AnalysisJob).where(AnalysisJob.id == job_id))
        job = result.scalar_one_or_none()
        if not job:
            return

        job.status = "running"
        job.started_at = datetime.now(timezone.utc)
        await db.commit()

        # Progress callback → WebSocket
        async def progress_callback(stage: str, progress: int, message: str = ""):
            job.stage = stage
            job.progress = progress
            await db.commit()
            # Push to WebSocket if connected
            ws = _ws_connections.get(job_id)
            if ws:
                try:
                    await ws.send_json({
                        "job_id": job_id,
                        "status": "running",
                        "stage": stage,
                        "progress": progress,
                        "message": message,
                    })
                except Exception:
                    pass

        try:
            pipeline = AnalysisPipeline(vlm_model=vlm_model)
            output = await pipeline.run(video_path, job_id, progress_callback)

            scores = output["scores"]

            # Persist result
            analysis_result = AnalysisResult(
                job_id=job_id,
                eye_contact_score=scores.get("eye_contact_score"),
                directed_expression_score=scores.get("directed_expression_score"),
                descriptive_gesture_score=scores.get("descriptive_gesture_score"),
                hand_mannerism_score=scores.get("hand_mannerism_score"),
                joint_attention_mean=scores.get("joint_attention_mean"),
                postural_orientation_mean=scores.get("postural_orientation_mean"),
                profile_deviations=output["profile_deviations"],
                feature_importance=output["feature_importance"],
                natural_language_explanation=output["natural_language_explanation"],
                vlm_frame_analyses=output["vlm_frame_analyses"],
                segment_timeline=output["segment_timeline"],
                frames_analysed=output["frames_analysed"],
                clinical_disclaimer=output["clinical_disclaimer"],
            )
            db.add(analysis_result)

            # Persist behaviour segments
            for seg in output.get("segment_timeline", []):
                segment = BehaviourSegment(
                    job_id=job_id,
                    start_time=seg.get("time", 0),
                    end_time=seg.get("time", 0) + 1.0,
                    behaviour_type=seg.get("type", "unknown"),
                    description=seg.get("event", ""),
                    extra_data=seg,
                )
                db.add(segment)

            job.status = "complete"
            job.progress = 100
            job.stage = "Complete"
            job.completed_at = datetime.now(timezone.utc)
            await db.commit()

            # Final WebSocket push
            ws = _ws_connections.get(job_id)
            if ws:
                try:
                    await ws.send_json({
                        "job_id": job_id,
                        "status": "complete",
                        "stage": "Complete",
                        "progress": 100,
                        "message": "Analysis complete",
                    })
                except Exception:
                    pass

        except Exception as e:
            logger.error("pipeline_failed", job_id=job_id, error=str(e))
            job.status = "failed"
            job.error_message = str(e)
            await db.commit()

            ws = _ws_connections.get(job_id)
            if ws:
                try:
                    await ws.send_json({
                        "job_id": job_id,
                        "status": "failed",
                        "stage": "Error",
                        "progress": job.progress,
                        "message": str(e),
                    })
                except Exception:
                    pass


@router.websocket("/ws/{job_id}")
async def analysis_progress_ws(
    websocket: WebSocket,
    job_id: str,
    token: str | None = None,
):
    """WebSocket endpoint for real-time pipeline progress. Auth via ?token= query param."""
    from app.core.security import decode_access_token
    from jose import JWTError
    if token:
        try:
            decode_access_token(token)
        except JWTError:
            await websocket.close(code=4001)
            return

    await websocket.accept()
    _ws_connections[job_id] = websocket
    try:
        # Send current status immediately on connect
        async with __import__("app.db.database", fromlist=["AsyncSessionLocal"]).AsyncSessionLocal() as db:
            result = await db.execute(select(AnalysisJob).where(AnalysisJob.id == job_id))
            job = result.scalar_one_or_none()
            if job:
                await websocket.send_json({
                    "job_id": job_id,
                    "status": job.status,
                    "stage": job.stage,
                    "progress": job.progress,
                    "message": "Connected",
                })
        # Keep connection alive
        while True:
            await asyncio.sleep(1)
            try:
                await websocket.send_json({"ping": True})
            except Exception:
                break
    except WebSocketDisconnect:
        pass
    finally:
        _ws_connections.pop(job_id, None)


@router.get("/jobs/{job_id}", response_model=AnalysisJobResponse)
async def get_job(job_id: str, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    result = await db.execute(select(AnalysisJob).where(AnalysisJob.id == job_id))
    job = result.scalar_one_or_none()
    if not job:
        raise HTTPException(status_code=404, detail="Analysis job not found")
    return AnalysisJobResponse.model_validate(job)


@router.get("/jobs", response_model=list[AnalysisJobResponse])
async def list_jobs(db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    # selectinload eager-loads the related VideoUpload for all jobs in one
    # extra query, instead of N+1 individual lookups per job in the list.
    result = await db.execute(
        select(AnalysisJob)
        .options(selectinload(AnalysisJob.video))
        .order_by(AnalysisJob.created_at.desc())
        .limit(20)
    )
    jobs = result.scalars().all()
    return [
        AnalysisJobResponse(
            id=j.id,
            video_id=j.video_id,
            video_filename=j.video.original_filename if j.video else None,
            status=j.status,
            stage=j.stage,
            progress=j.progress,
            error_message=j.error_message,
            vlm_model=j.vlm_model,
            created_at=j.created_at,
            started_at=j.started_at,
            completed_at=j.completed_at,
        )
        for j in jobs
    ]


@router.get("/results/{job_id}", response_model=AnalysisResultResponse)
async def get_result(job_id: str, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    result = await db.execute(
        select(AnalysisResult).where(AnalysisResult.job_id == job_id)
    )
    r = result.scalar_one_or_none()
    if not r:
        raise HTTPException(status_code=404, detail="Results not yet available")

    # Look up the original video filename for display/reference — same join
    # pattern used elsewhere in this file (AnalysisJob by id, then .video).
    # This is the human-readable identifier (e.g. "FRIAM07_side.mp4") for
    # matching a result against participant records; job_id/video_id are
    # opaque UUIDs not useful for that on their own.
    video_filename = None
    job_result = await db.execute(select(AnalysisJob).where(AnalysisJob.id == r.job_id))
    job = job_result.scalar_one_or_none()
    if job:
        video_result = await db.execute(select(VideoUpload).where(VideoUpload.id == job.video_id))
        video = video_result.scalar_one_or_none()
        if video:
            video_filename = video.original_filename

    return AnalysisResultResponse(
        id=r.id,
        job_id=r.job_id,
        video_filename=video_filename,
        ados_scores=ADOSCoreScores(
            eye_contact_score=r.eye_contact_score,
            directed_expression_score=r.directed_expression_score,
            descriptive_gesture_score=r.descriptive_gesture_score,
            hand_mannerism_score=r.hand_mannerism_score,
        ),
        hri_extensions=HRIExtensionScores(
            joint_attention_mean=r.joint_attention_mean,
            postural_orientation_mean=r.postural_orientation_mean,
        ),
        profile_summary=ProfileSummary(
            profile_deviations=r.profile_deviations,
            feature_importance=r.feature_importance,
        ),
        natural_language_explanation=r.natural_language_explanation,
        segment_timeline=r.segment_timeline,
        frames_analysed=r.frames_analysed,
        clinical_disclaimer=r.clinical_disclaimer,
        created_at=r.created_at,
    )


# ─── Frame Summaries & Export Endpoints ─────────────────────────────────────

@router.get("/results/{job_id}/frame-summaries", response_model=FrameSummariesResponse)
async def get_frame_summaries(
    job_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Get frame-by-frame VLM summaries for a specific job.
    This shows what the VLM said about each frame in the video.
    Returns full raw_response with NO truncation.
    """
    # Get the analysis result
    result = await db.execute(
        select(AnalysisResult).where(AnalysisResult.job_id == job_id)
    )
    r = result.scalar_one_or_none()
    if not r:
        raise HTTPException(status_code=404, detail="Results not found for this job")
    
    # Get job info
    job_result = await db.execute(
        select(AnalysisJob).where(AnalysisJob.id == job_id)
    )
    job = job_result.scalar_one_or_none()
    
    # Parse frame analyses with robust error handling
    summaries = []
    if r.vlm_frame_analyses:
        try:
            # Handle different data types
            if isinstance(r.vlm_frame_analyses, list):
                frames = r.vlm_frame_analyses
            elif isinstance(r.vlm_frame_analyses, str):
                # Try JSON first, then ast.literal_eval
                try:
                    frames = json.loads(r.vlm_frame_analyses)
                except json.JSONDecodeError:
                    try:
                        frames = ast.literal_eval(r.vlm_frame_analyses)
                    except:
                        # If all fails, treat as single string
                        frames = [{"summary": str(r.vlm_frame_analyses), "timestamp": 0}]
            else:
                frames = []
            
            # Ensure frames is a list
            if not isinstance(frames, list):
                frames = []
            
            for i, frame in enumerate(frames):
                # Handle both dict and string formats
                if isinstance(frame, str):
                    try:
                        frame_data = json.loads(frame)
                    except:
                        frame_data = {"summary": frame, "timestamp": 0, "eval_count": 0, "parse_error": False}
                elif isinstance(frame, dict):
                    frame_data = frame
                else:
                    frame_data = {"summary": str(frame), "timestamp": 0, "eval_count": 0, "parse_error": False}
                
                # ✅ FULL raw response - NO TRUNCATION
                raw_response = frame_data.get("raw_response", "")
                
                summaries.append(FrameSummaryResponse(
                    frame_index=i + 1,
                    timestamp=frame_data.get("timestamp", 0),
                    summary=frame_data.get("summary", "No summary available"),
                    tokens=frame_data.get("eval_count", 0),
                    parse_error=frame_data.get("parse_error", False),
                    raw_response=raw_response,  # Full response
                ))
                
        except Exception as e:
            logger.error(f"Error parsing frame summaries for job {job_id}: {e}")
            # Return empty list instead of failing
            pass
    
    return FrameSummariesResponse(
        job_id=job_id,
        video_id=job.video_id if job else None,
        vlm_model=job.vlm_model if job else None,
        frames_analysed=r.frames_analysed or len(summaries),
        frame_summaries=summaries,
        natural_language_explanation=r.natural_language_explanation,
    )


@router.get("/results/{job_id}/export")
async def export_results(
    job_id: str,
    format: str = "json",
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Export analysis results as JSON or Excel file.
    Includes frame-by-frame VLM summaries with full raw responses (NO truncation).

    JSON Export: Full structured data with all fields and complete raw responses.

    Excel Export (6 sheets):
        1. Summary - Job metadata
        2. Frame Summaries - All frames with summaries and raw responses
        3. ADOS Scores - ADOS core items (4 items; Items 3 and 5 removed, see below)
        4. HRI Metrics - HRI extension metrics (2 items; Multimodal Sync Ratio removed)
        5. Profile Summary - Deviation-from-typical ranking (replaces AQ-10/SHAP)
        6. Explanation - Natural language explanation
    """
    # Get the analysis result
    result = await db.execute(
        select(AnalysisResult).where(AnalysisResult.job_id == job_id)
    )
    r = result.scalar_one_or_none()
    if not r:
        raise HTTPException(status_code=404, detail="Results not found")
    
    # Get job info
    job_result = await db.execute(
        select(AnalysisJob).where(AnalysisJob.id == job_id)
    )
    job = job_result.scalar_one_or_none()
    
    # Build export data
    export_data = {
        "job_id": job_id,
        "video_id": job.video_id if job else None,
        "vlm_model": job.vlm_model if job else None,
        "created_at": r.created_at.isoformat() if r.created_at else None,
        "frames_analysed": r.frames_analysed,
        "ados_scores": {
            "eye_contact": r.eye_contact_score,
            "directed_expression": r.directed_expression_score,
            "descriptive_gesture": r.descriptive_gesture_score,
            "hand_mannerism": r.hand_mannerism_score,
        },
        "hri_metrics": {
            "joint_attention_mean": r.joint_attention_mean,
            "postural_orientation_mean": r.postural_orientation_mean,
        },
        "profile_summary": r.feature_importance if r.feature_importance else [],
        "natural_language_explanation": r.natural_language_explanation,
        "clinical_disclaimer": r.clinical_disclaimer,
        "frame_summaries": []
    }
    
    # Get frame summaries with FULL raw responses (NO truncation)
    if r.vlm_frame_analyses:
        try:
            # Parse the data
            if isinstance(r.vlm_frame_analyses, list):
                frames = r.vlm_frame_analyses
            elif isinstance(r.vlm_frame_analyses, str):
                try:
                    frames = json.loads(r.vlm_frame_analyses)
                except:
                    frames = []
            else:
                frames = []
            
            for i, frame in enumerate(frames):
                if isinstance(frame, dict):
                    export_data["frame_summaries"].append({
                        "frame_index": i + 1,
                        "timestamp": frame.get("timestamp", 0),
                        "summary": frame.get("summary", "No summary"),
                        "tokens": frame.get("eval_count", 0),
                        "parse_error": frame.get("parse_error", False),
                        "raw_response": frame.get("raw_response", ""),  # FULL response
                        "consistency_flags": frame.get("consistency_flags", []),
                    })
                elif isinstance(frame, str):
                    export_data["frame_summaries"].append({
                        "frame_index": i + 1,
                        "timestamp": 0,
                        "summary": frame[:200] if frame else "Empty frame",
                        "tokens": 0,
                        "parse_error": False,
                        "raw_response": frame,
                    })
        except Exception as e:
            logger.error(f"Error building frame summaries for export: {e}")
    
    # Export based on format
    if format in ("excel", "xlsx"):
        try:
            import pandas as pd
            output = BytesIO()

            # Data-quality figure: what fraction of frames actually produced
            # usable VLM data vs fell back (parse failure, timeout, or —
            # after the template-leak fix — echoed placeholder text). This is
            # computed the same way regardless of *why* a frame failed, so it
            # stays meaningful as new failure modes are caught over time.
            total_frames = len(export_data["frame_summaries"])
            failed_frames = sum(
                1 for f in export_data["frame_summaries"] if f.get("parse_error")
            )
            valid_frames = total_frames - failed_frames
            valid_pct = round(100 * valid_frames / total_frames, 1) if total_frames else 0.0
            data_quality_warning = (
                f"WARNING: only {valid_pct}% of frames yielded usable VLM data — "
                f"treat ADOS/HRI results below as unreliable for this job."
                if valid_pct < 50 else ""
            )

            top_indicator = r.feature_importance[0] if r.feature_importance else None

            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Sheet 1: Summary
                summary_df = pd.DataFrame([{
                    "Job ID": job_id,
                    "Video ID": job.video_id if job else None,
                    "Model": job.vlm_model if job else None,
                    "Frames Analysed": r.frames_analysed,
                    "Valid Frames (%)": valid_pct,
                    "Most Notable Indicator": top_indicator["label"] if top_indicator else None,
                    "Notability": top_indicator["tier"] if top_indicator else None,
                    "Created At": r.created_at.isoformat() if r.created_at else None,
                    "Data Quality Warning": data_quality_warning,
                }])
                summary_df.to_excel(writer, sheet_name="Summary", index=False)
                ws_sum = writer.sheets["Summary"]
                ws_sum.column_dimensions["A"].width = 28
                ws_sum.column_dimensions["B"].width = 40
                ws_sum.column_dimensions["I"].width = 70
                if data_quality_warning:
                    from openpyxl.styles import PatternFill as _PatternFill, Font as _Font
                    warn_fill = _PatternFill("solid", fgColor="C0392B")
                    warn_font = _Font(color="FFFFFF", bold=True)
                    for cell in ws_sum[2]:
                        cell.fill = warn_fill
                        cell.font = warn_font
                
                # Sheet 2: Frame Summaries (FULL raw responses)
                if export_data["frame_summaries"]:
                    # Serialise raw_response to readable JSON string for Excel
                    frames_for_excel = []
                    for frame in export_data["frame_summaries"]:
                        row = dict(frame)
                        raw = row.get("raw_response", "")
                        if isinstance(raw, str) and raw.strip():
                            clean = raw.strip()
                            if "```" in clean:
                                for part in clean.split("```"):
                                    part = part.strip()
                                    if part.startswith("json"):
                                        part = part[4:].strip()
                                    if part.startswith("{"):
                                        clean = part
                                        break
                            start = clean.find("{")
                            end   = clean.rfind("}") + 1
                            if start >= 0 and end > start:
                                try:
                                    row["raw_response"] = json.dumps(
                                        json.loads(clean[start:end]), indent=2
                                    )
                                except Exception:
                                    row["raw_response"] = raw
                        # Flatten consistency_flags list to a readable string —
                        # a raw Python list doesn't render cleanly as an Excel cell.
                        flags = row.get("consistency_flags", [])
                        row["consistency_flags"] = ", ".join(flags) if flags else ""
                        frames_for_excel.append(row)

                    frames_df = pd.DataFrame(frames_for_excel)
                    frames_df.to_excel(writer, sheet_name="Frame Summaries", index=False)

                    # Auto-size columns and wrap text for readability
                    ws = writer.sheets["Frame Summaries"]
                    col_widths = {
                        "A": 12,  # frame_index
                        "B": 12,  # timestamp
                        "C": 60,  # summary
                        "D": 10,  # tokens
                        "E": 12,  # parse_error
                        "F": 80,  # raw_response
                        "G": 45,  # consistency_flags
                    }
                    for col_letter, width in col_widths.items():
                        ws.column_dimensions[col_letter].width = width
                    from openpyxl.styles import Alignment, PatternFill, Font
                    header_fill = PatternFill("solid", fgColor="2E4057")
                    header_font = Font(color="FFFFFF", bold=True)
                    for cell in ws[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                    wrap_align = Alignment(wrap_text=True, vertical="top")
                    for row in ws.iter_rows(min_row=2):
                        for cell in row:
                            cell.alignment = wrap_align
                
                # Sheet 3: ADOS Scores
                ados_labels = {
                    "eye_contact": "Eye Contact (Item 1)",
                    "directed_expression": "Directed Expressions (Item 2)",
                    "descriptive_gesture": "Descriptive Gestures (Item 4)",
                    "hand_mannerism": "Hand & Finger Mannerisms (Item 6)",
                }
                ados_data = {ados_labels.get(k, k): v for k, v in export_data["ados_scores"].items()}
                ados_df = pd.DataFrame([ados_data])
                ados_df.to_excel(writer, sheet_name="ADOS Scores", index=False)
                ws_ados = writer.sheets["ADOS Scores"]
                for col in ws_ados.columns:
                    ws_ados.column_dimensions[col[0].column_letter].width = max(
                        len(str(col[0].value or "")), 12
                    )
                
                # Sheet 4: HRI Metrics
                hri_labels = {
                    "joint_attention_mean": "Joint Attention Mean (0-1)",
                    "postural_orientation_mean": "Postural Orientation Mean (0-1)",
                }
                hri_data = {hri_labels.get(k, k): v for k, v in export_data["hri_metrics"].items()}
                hri_df = pd.DataFrame([hri_data])
                hri_df.to_excel(writer, sheet_name="HRI Metrics", index=False)
                ws_hri = writer.sheets["HRI Metrics"]
                for col in ws_hri.columns:
                    ws_hri.column_dimensions[col[0].column_letter].width = max(
                        len(str(col[0].value or "")), 14
                    )
                
                # Sheet 5: Profile Summary — replaces SHAP Values. No published,
                # validated formula exists for predicting AQ-10 from visual-only
                # indicators (confirmed with supervisor), so this sheet explains
                # the ADOS/HRI profile directly: each indicator's deviation from
                # typical/neutral, ranked most-notable first.
                if r.feature_importance:
                    profile_df = pd.DataFrame(r.feature_importance)
                    profile_df.to_excel(writer, sheet_name="Profile Summary", index=False)
                    ws_profile = writer.sheets["Profile Summary"]
                    for col in ws_profile.columns:
                        ws_profile.column_dimensions[col[0].column_letter].width = max(
                            len(str(col[0].value or "")), 14
                        )
                
                # Sheet 6: Explanation
                if r.natural_language_explanation:
                    explanation_text = (
                        r.natural_language_explanation
                        .replace("\u2193", "↓")
                        .replace("\u2191", "↑")
                        .replace("\u26a0", "⚠")
                        .replace("**", "")
                    )
                    expl_df = pd.DataFrame([{"Explanation": explanation_text}])
                    expl_df.to_excel(writer, sheet_name="Explanation", index=False)
                    ws_expl = writer.sheets["Explanation"]
                    ws_expl.column_dimensions["A"].width = 120
                    ws_expl.row_dimensions[2].height = 200
                    from openpyxl.styles import Alignment as _Align
                    if ws_expl["A2"].value:
                        ws_expl["A2"].alignment = _Align(wrap_text=True, vertical="top")
            
            output.seek(0)
            filename = f"hri_results_{job_id[:8]}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
        except ImportError:
            raise HTTPException(
                status_code=500, 
                detail="pandas and openpyxl are required for Excel export. Install with: pip install pandas openpyxl"
            )
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Excel export failed: {str(e)}")
    
    else:
        # JSON format — pretty-printed, raw_response decoded to object where possible
        filename = f"hri_results_{job_id[:8]}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"

        # Decode raw_response strings into parsed JSON objects for readability
        for frame in export_data.get("frame_summaries", []):
            raw = frame.get("raw_response", "")
            if isinstance(raw, str) and raw.strip():
                clean = raw.strip()
                # Strip markdown fences if present
                if "```" in clean:
                    for part in clean.split("```"):
                        part = part.strip()
                        if part.startswith("json"):
                            part = part[4:].strip()
                        if part.startswith("{"):
                            clean = part
                            break
                start = clean.find("{")
                end   = clean.rfind("}") + 1
                if start >= 0 and end > start:
                    try:
                        frame["raw_response"] = json.loads(clean[start:end])
                    except json.JSONDecodeError:
                        pass  # keep as string if parse fails

        # Decode natural_language_explanation unicode escapes
        if isinstance(export_data.get("natural_language_explanation"), str):
            export_data["natural_language_explanation"] = (
                export_data["natural_language_explanation"]
                .replace("\u2193", "↓")
                .replace("\u2191", "↑")
                .replace("\u26a0", "⚠")
            )

        json_bytes = json.dumps(export_data, indent=2, ensure_ascii=False).encode("utf-8")

        return StreamingResponse(
            BytesIO(json_bytes),
            media_type="application/json",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )